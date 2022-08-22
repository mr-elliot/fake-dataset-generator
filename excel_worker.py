import math
import random
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook

global wb
global ws
wb = Workbook()
ws = wb.active


def words_generator():
    # returns 75.8K words in total
    url = "https://www.mit.edu/~ecprice/wordlist.10000"
    html = urlopen(url).read()
    soup = BeautifulSoup(html, features="html.parser")
    text = list(map(str, soup.get_text().split('\n')))
    return text


def adding_serial_number(n):
    # Executes when a user asks for the serial number
    ws.cell(1, 1).value = "Serial Number"
    for i in range(1, n):
        ws.cell(i+1, 1).value = i


def add_noise(row, col, sl, noise_level):
    """
        we are fixing the column and picking the row number randomly to remove corresponding value.
        row   = 1(n) + random_number
        row'  = row + random_number
        row'' = row' + random_number
    """

    till = 2 if row <= 5 else math.ceil(row / int(noise_level))
    start = 2 if sl else 1

    for c in range(start, col+2):
        n = 1
        while n < (row - till):
            random_number = random.randint(1, till)
            pos = n + random_number
            ws.cell(pos, c).value = None
            n = pos + 1


def creating_excel(nrow, ncol, col_names, col_features, noise, noise_lvl, sl_no, file_name):

    if 'string' in col_features:
        rand_words = words_generator()

    start_point = 2 if sl_no else 1
    for col, *arr in enumerate(zip(col_names, col_features), start_point):
        name, feature = arr[0]
        row = 1
        ws.cell(row, col).value = name      # giving the name of the column

        if feature == 'integer':
            row += 1
            temp = (nrow - (nrow//4))       # to get range till 75%
            while row <= nrow:
                val = random.randint(1, temp)
                ws.cell(row, col).value = val
                row += 1

        elif feature == 'float':
            row += 1
            temp = (nrow - (nrow // 4))
            while row <= nrow:
                val = round(random.uniform(1, temp), 3)
                ws.cell(row, col).value = val
                row += 1

        elif feature == 'string':
            row += 1
            while row <= nrow:
                val = rand_words[random.randint(1, 10000)]
                ws.cell(row, col).value = val
                row += 1

    if noise:
        add_noise(nrow, ncol, sl_no, noise_lvl)

    wb.save(filename=file_name + ".xlsx")
    print("\n\nIf you trying to generate 1000's of row keep the noise percent in 100's")
    return






